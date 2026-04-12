#!/usr/bin/env python3
"""
CYBERSHEPPARD - Compliance Data Population Script

Populates compliance_controls table with all 113 controls from Excel file.
Production-ready script with proper error handling and transaction management.
"""

import openpyxl
import psycopg2
import psycopg2.extras
import sys
import os
from typing import List, Dict, Any

# Database connection parameters (from environment or defaults)
DB_CONFIG = {
    'host': os.getenv('DB_HOST', 'localhost'),
    'port': int(os.getenv('DB_PORT', '5432')),
    'database': os.getenv('DB_NAME', 'cybersheppard'),
    'user': os.getenv('DB_USER', 'cybersheppard'),
    'password': os.getenv('DB_PASSWORD', 'cybersheppard')
}

def parse_excel_controls(filepath: str) -> List[Dict[str, Any]]:
    """
    Parse Excel file and extract all controls with their metadata.

    Returns:
        List of control dictionaries
    """
    print(f"📖 Reading Excel file: {filepath}")

    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb['Master Mapping']

    # Get headers from row 1
    headers = []
    for cell in ws[1]:
        headers.append(cell.value)

    controls = []

    # Parse all data rows
    for row_idx in range(2, ws.max_row + 1):
        row = ws[row_idx]
        control_raw = {}

        for col_idx, cell in enumerate(row):
            if col_idx < len(headers) and headers[col_idx]:
                control_raw[headers[col_idx]] = cell.value

        # Only process rows with valid requirement
        if not control_raw.get('Host Requirement'):
            continue

        # Transform to database-friendly format
        control = transform_control(control_raw)
        controls.append(control)

    print(f"✅ Parsed {len(controls)} controls from Excel")
    return controls

def transform_control(raw: Dict[str, Any]) -> Dict[str, Any]:
    """
    Transform raw Excel data to database-ready format.
    """
    # Parse framework references into arrays
    nis2_refs = parse_references(raw.get('NIS2 Reference'))
    nist_refs = parse_references(raw.get('NIST 800-53 Rev5'))
    iso_refs = parse_references(raw.get('ISO 27001:2022'))
    mitre_refs = parse_references(raw.get('MITRE D3FEND'))

    # Determine framework applicability
    applies_to_nis2 = bool(nis2_refs)
    applies_to_nist = bool(nist_refs)
    applies_to_iso = bool(iso_refs)
    applies_to_mitre = bool(mitre_refs)
    applies_to_all = applies_to_nis2 and applies_to_nist and applies_to_iso and applies_to_mitre

    return {
        'macroarea': raw.get('Macroarea'),
        'sub_control': raw.get('Sub-Control'),
        'sub_sub_control': raw.get('Sub-Sub-Control'),
        'requirement': raw.get('Host Requirement'),
        'priority': raw.get('Priority', 'Medium'),
        'implementation_notes': raw.get('Implementation Notes'),
        'nis2_references': nis2_refs,
        'nist_references': nist_refs,
        'iso_references': iso_refs,
        'mitre_references': mitre_refs,
        'applies_to_nis2': applies_to_nis2,
        'applies_to_nist': applies_to_nist,
        'applies_to_iso': applies_to_iso,
        'applies_to_mitre': applies_to_mitre,
        'applies_to_all_frameworks': applies_to_all
    }

def parse_references(ref_string: str) -> List[str]:
    """
    Parse comma-separated reference string into list.

    Examples:
        "IA-2(1), IA-2(2), IA-2(12)" -> ['IA-2(1)', 'IA-2(2)', 'IA-2(12)']
        "A.5.17, A.5.18" -> ['A.5.17', 'A.5.18']
    """
    if not ref_string or ref_string == 'N/A':
        return []

    # Split by comma and clean whitespace
    refs = [r.strip() for r in str(ref_string).split(',')]
    return [r for r in refs if r and r != 'N/A']

def get_macroarea_id(conn, macroarea_name: str) -> int:
    """Get macroarea ID by name."""
    with conn.cursor() as cur:
        cur.execute(
            "SELECT id FROM compliance_macroareas WHERE name = %s",
            (macroarea_name,)
        )
        result = cur.fetchone()
        if not result:
            raise ValueError(f"Macroarea not found: {macroarea_name}")
        return result[0]

def parse_os_implementations(filepath: str) -> Dict[str, Dict[str, bool]]:
    """
    Parse OS Implementation Matrix to get OS support for each requirement.

    Returns:
        Dict mapping requirement -> {os_name: supported}
    """
    print(f"📖 Reading OS Implementation Matrix...")

    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb['OS Implementation Matrix']

    # Get headers
    headers = []
    for cell in ws[1]:
        headers.append(cell.value)

    implementations = {}

    for row_idx in range(2, ws.max_row + 1):
        row = ws[row_idx]
        requirement = None
        os_support = {}

        for col_idx, cell in enumerate(row):
            if col_idx < len(headers):
                header = headers[col_idx]
                value = cell.value

                if header == 'Requirement':
                    requirement = value
                elif header in ['Debian/Ubuntu', 'RHEL/Oracle', 'SLES', 'Windows 2019', 'Windows 2022', 'Docker', 'LXC']:
                    # Check if implemented (contains "Implemented" or checkmark)
                    is_implemented = value and ('Implemented' in str(value) or '✓' in str(value))

                    # Map to database column names
                    os_map = {
                        'Debian/Ubuntu': 'debian_ubuntu',
                        'RHEL/Oracle': 'rhel_oracle',
                        'SLES': 'sles',
                        'Windows 2019': 'windows_2019',
                        'Windows 2022': 'windows_2022',
                        'Docker': 'docker',
                        'LXC': 'lxc'
                    }
                    os_support[os_map[header]] = is_implemented

        if requirement and os_support:
            implementations[requirement] = os_support

    print(f"✅ Parsed OS support for {len(implementations)} requirements")
    return implementations

def insert_controls(conn, controls: List[Dict[str, Any]], os_implementations: Dict[str, Dict[str, bool]]):
    """
    Insert all controls into database with OS support information.
    """
    print(f"\n📝 Inserting {len(controls)} controls into database...")

    inserted = 0
    skipped = 0

    with conn.cursor() as cur:
        for control in controls:
            try:
                # Get macroarea ID
                macroarea_id = get_macroarea_id(conn, control['macroarea'])

                # Get OS support for this requirement
                os_support = os_implementations.get(control['requirement'], {})

                # Insert control
                cur.execute("""
                    INSERT INTO compliance_controls (
                        macroarea_id,
                        sub_control,
                        sub_sub_control,
                        requirement,
                        priority,
                        implementation_notes,
                        nis2_references,
                        nist_references,
                        iso_references,
                        mitre_references,
                        applies_to_nis2,
                        applies_to_nist,
                        applies_to_iso,
                        applies_to_mitre,
                        applies_to_all_frameworks,
                        supports_debian_ubuntu,
                        supports_rhel_oracle,
                        supports_sles,
                        supports_windows_2019,
                        supports_windows_2022,
                        supports_docker,
                        supports_lxc
                    ) VALUES (
                        %s, %s, %s, %s, %s, %s, %s, %s, %s, %s,
                        %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s
                    )
                    ON CONFLICT (macroarea_id, requirement) DO NOTHING
                """, (
                    macroarea_id,
                    control['sub_control'],
                    control['sub_sub_control'],
                    control['requirement'],
                    control['priority'],
                    control['implementation_notes'],
                    control['nis2_references'],
                    control['nist_references'],
                    control['iso_references'],
                    control['mitre_references'],
                    control['applies_to_nis2'],
                    control['applies_to_nist'],
                    control['applies_to_iso'],
                    control['applies_to_mitre'],
                    control['applies_to_all_frameworks'],
                    os_support.get('debian_ubuntu', False),
                    os_support.get('rhel_oracle', False),
                    os_support.get('sles', False),
                    os_support.get('windows_2019', False),
                    os_support.get('windows_2022', False),
                    os_support.get('docker', False),
                    os_support.get('lxc', False)
                ))

                if cur.rowcount > 0:
                    inserted += 1
                else:
                    skipped += 1

            except Exception as e:
                print(f"❌ Error inserting control '{control.get('requirement', 'unknown')}': {e}")
                raise

    print(f"✅ Inserted: {inserted}, Skipped (already exist): {skipped}")

def verify_data(conn):
    """Verify inserted data."""
    print(f"\n🔍 Verifying inserted data...")

    with conn.cursor(cursor_factory=psycopg2.extras.DictCursor) as cur:
        # Count controls
        cur.execute("SELECT COUNT(*) FROM compliance_controls")
        total_controls = cur.fetchone()[0]
        print(f"   Total controls in database: {total_controls}")

        # Count by macroarea
        cur.execute("""
            SELECT m.name, COUNT(c.id)
            FROM compliance_macroareas m
            LEFT JOIN compliance_controls c ON c.macroarea_id = m.id
            GROUP BY m.name
            ORDER BY COUNT(c.id) DESC
        """)
        print(f"\n   Controls per macroarea:")
        for row in cur.fetchall():
            print(f"     - {row[0]}: {row[1]} controls")

        # Count by framework
        cur.execute("""
            SELECT
                COUNT(*) FILTER (WHERE applies_to_nis2) as nis2,
                COUNT(*) FILTER (WHERE applies_to_nist) as nist,
                COUNT(*) FILTER (WHERE applies_to_iso) as iso,
                COUNT(*) FILTER (WHERE applies_to_mitre) as mitre,
                COUNT(*) FILTER (WHERE applies_to_all_frameworks) as all_four
            FROM compliance_controls
        """)
        row = cur.fetchone()
        print(f"\n   Controls per framework:")
        print(f"     - NIS2: {row[0]} controls")
        print(f"     - NIST 800-53: {row[1]} controls")
        print(f"     - ISO 27001: {row[2]} controls")
        print(f"     - MITRE D3FEND: {row[3]} controls")
        print(f"     - All 4 frameworks: {row[4]} controls")

        # Count by priority
        cur.execute("""
            SELECT priority, COUNT(*)
            FROM compliance_controls
            GROUP BY priority
            ORDER BY
                CASE priority
                    WHEN 'Critical' THEN 1
                    WHEN 'High' THEN 2
                    WHEN 'Medium' THEN 3
                    WHEN 'Low' THEN 4
                END
        """)
        print(f"\n   Controls by priority:")
        for row in cur.fetchall():
            print(f"     - {row[0]}: {row[1]} controls")

def main():
    """Main execution flow."""
    excel_filepath = 'documentazione/Host_Compliance_Framework_Mapping(4).xlsx'

    if not os.path.exists(excel_filepath):
        print(f"❌ Excel file not found: {excel_filepath}")
        return 1

    try:
        # Parse Excel
        controls = parse_excel_controls(excel_filepath)
        os_implementations = parse_os_implementations(excel_filepath)

        # Connect to database
        print(f"\n🔌 Connecting to database: {DB_CONFIG['host']}:{DB_CONFIG['port']}/{DB_CONFIG['database']}")
        conn = psycopg2.connect(**DB_CONFIG)
        conn.autocommit = False

        try:
            # Insert data
            insert_controls(conn, controls, os_implementations)

            # Verify
            verify_data(conn)

            # Commit transaction
            conn.commit()
            print(f"\n✅ All data committed successfully!")

        except Exception as e:
            conn.rollback()
            print(f"\n❌ Error during database operations, rolling back: {e}")
            raise
        finally:
            conn.close()

        return 0

    except Exception as e:
        print(f"\n❌ Fatal error: {e}")
        import traceback
        traceback.print_exc()
        return 1

if __name__ == '__main__':
    sys.exit(main())
